import openpyxl
import random
import os
from tkinter import *
from tkinter import filedialog, messagebox
import ctypes
import platform
import tkinter as tk

file = ''

# 设置高DPI感知
def set_dpi_awareness():
    if platform.system() == 'Windows':
        try:
            # 检查Windows版本
            if int(platform.release()) >= 8:
                # 适用于 Windows 8.1 及以上版本
                ctypes.windll.shcore.SetProcessDpiAwareness(1)
            else:
                # 适用于 Windows 7 及更低版本
                ctypes.windll.user32.SetProcessDPIAware()
        except Exception as e:
            print("DPI awareness setting failed:", e)

# 设置 DPI 感知
set_dpi_awareness()


# 文件选择函数
def select_file():
    global file
    file = filedialog.askopenfilename(title="请选择Excel文件", filetypes=[("Excel文件", "*.xlsx")])
    Entry1.delete(0, END)
    Entry1.insert(0, file)

# 随机排列座位函数
def random_seat_allocation(file, rows, cols):
    if file == '':
        messagebox.showinfo("提示", "请先选择文件")
        return
    
    # 打开文件并获取名单
    wb = openpyxl.load_workbook(file)
    sheet_name = wb['名单'] # 找到名单的sheet
    names = []

    # 读取名单
    for i in range(1, sheet_name.max_row):
        name = sheet_name.cell(i + 1, 2).value
        if name:
            names.append(name)

    # 检查名单是否超过座位数
    if len(names) > rows * cols:
        messagebox.showinfo("提示", "座位不够，请增加行数或列数")
        return

    # 随机打乱名单
    random.shuffle(names)

    # 创建新的工作表
    new_sheet = wb.create_sheet("随机排列后的名字")

    # 填充表格
    name_index = 0
    for i in range(rows):
        for j in range(cols):
            if name_index < len(names):
                new_sheet.cell(i + 1, j + 1).value = names[name_index]
                name_index += 1

    # 保存文件
    wb.save(file)
    messagebox.showinfo("提示", "座位随机排列完成并已保存到文件中")

    # 弹窗询问是否打开文件、文件夹或者关闭程序
    user_choice = messagebox.askyesnocancel("操作", "是否要打开文件？")
    
    if user_choice:  # 用户选择打开文件
        os.startfile(file)  # 打开文件
    elif user_choice is None:  # 用户选择取消
        return
    else:  # 用户选择打开文件夹
        folder_path = os.path.dirname(file)  # 获取文件所在文件夹路径
        os.startfile(folder_path)  # 打开文件夹

# 排座位函数，获取输入并调用随机排列函数
def seat():
    try:
        rows = int(Entry2.get())
        cols = int(Entry3.get())
        random_seat_allocation(file, rows, cols)
    except ValueError:
        messagebox.showinfo("提示", "请输入有效的行数和列数")

# 主函数，创建界面
def main():
    global Entry1, Entry2, Entry3
    root = Tk()
    root.title("学生随机排座位软件v1.0")
    root.geometry("900x400")  # 设置窗口大小
    root.configure(bg="#F0F0F0")  # 设置背景颜色

    # 标题
    Label(root, text="学生随机排座位软件v1.0", font=("微软雅黑", 24, "bold"), bg="#F0F0F0", fg="#333333").grid(row=0, column=0, columnspan=3, pady=20)

    # 文件选择部分
    Label(root, text="请选择文件：", font=("Helvetica", 14), bg="#F0F0F0", fg="#333333").grid(row=1, column=0, sticky=E, padx=(20, 5))
    Entry1 = Entry(root, width=40, font=("Helvetica", 12))
    Entry1.grid(row=1, column=1, padx=5)
    Button(root, text="选择文件", font=("Helvetica", 12), command=select_file, bg="#4CAF50", fg="white", relief=FLAT, cursor="hand2").grid(row=1, column=2, padx=(5, 20))

    # 行数输入
    Label(root, text="请输入行数：", font=("Helvetica", 14), bg="#F0F0F0", fg="#333333").grid(row=2, column=0, sticky=E, padx=(20, 5), pady=(10, 0))
    Entry2 = Entry(root, width=10, font=("Helvetica", 12))
    Entry2.grid(row=2, column=1, sticky=W, padx=5, pady=(10, 0))

    # 列数输入
    Label(root, text="请输入列数：", font=("Helvetica", 14), bg="#F0F0F0", fg="#333333").grid(row=3, column=0, sticky=E, padx=(20, 5), pady=(10, 0))
    Entry3 = Entry(root, width=10, font=("Helvetica", 12))
    Entry3.grid(row=3, column=1, sticky=W, padx=5, pady=(10, 0))

    # 开始按钮
    Button(root, text="开始排座位", font=("Helvetica", 14), command=seat, bg="#2196F3", fg="white", relief=FLAT, cursor="hand2").grid(row=4, column=0, columnspan=3, pady=30)

    root.mainloop()


# 运行程序
if __name__ == "__main__":
    main()
